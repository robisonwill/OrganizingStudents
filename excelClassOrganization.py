# Harley Sigmon, Christian Yoder, Hunter Johanson, Paris Ward, Will Robison
# This program takes student data from an excel workbook and creates a new workbook with organized data
# It also calculates different measures based on the grades for each class

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

importedWorkbook = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")

def summaryInfo(ws):
    ws["F1"] = "Summary Statistics"
    bold_font = Font(bold=True)
    ws["F1"].font = bold_font

    ws["G1"] = "Value"
    bold_font = Font(bold=True)
    ws["G1"].font = bold_font

    ws["F2"] = "Highest Grade"
    ws["G2"] = "=MAX(D:D)"

    ws["F3"] = "Lowest Grade"
    ws["G3"] = "=MIN(D:D)"

    ws["F4"] = "Mean Grade"
    ws["G4"] = "=AVERAGE(D:D)"

    ws["F5"] = "Median Grade"
    ws["G5"] = "=MEDIAN(D:D)"

    ws["F6"] = "Number of Students"
    ws["G6"] = '=COUNTA(A:A)-1'

    cells_to_autofit = ["F1", "F2", "F3", "F4", "F5", "F6", "G1", "G2", "G3", "G4", "G5", "G6"]

    # Autofit logic
    for cell in cells_to_autofit:
        col_letter = cell[0]  # Extract column letter (e.g., "F" from "F1")
        cell_value = ws[cell].value
        if cell_value:
            current_width = ws.column_dimensions[col_letter].width or 0
            ws.column_dimensions[col_letter].width = max(current_width, len(str(cell_value)) + 2)  # Padding

currSheet = importedWorkbook.active

createWorkbook = Workbook()


# Define the header
header = ["Last Name", "First Name", "Student ID", "Grade"]


for row in currSheet.iter_rows(min_row= 2, min_col= 1, max_col= 1) :
    for cell in row :
        if cell.value not in createWorkbook.sheetnames :
            createWorkbook.create_sheet(f"{cell.value}")
        else:
            pass
createWorkbook.remove(createWorkbook["Sheet"])

# add the header to each sheet
for sheet in createWorkbook.sheetnames:
    createWorkbook[sheet].append(header)

# Creates list of student data from unorganized sheet
studentList = []
for row in currSheet.iter_rows(min_row=2, values_only=True):
    studentList.append(row)

# Iterates through student list and adds their data to the corresponding sheet
for student in studentList:
    currSheet = createWorkbook[student[0]]
    student_info = student[1].split("_")
    student_info.append(student[2])
    currSheet.append(student_info)

for sheet_name in createWorkbook.sheetnames:
    sheet = createWorkbook[sheet_name]

    summaryInfo(sheet)
    
    # Find the last row dynamically
    last_row = sheet.max_row  

    # Apply filter
    sheet.auto_filter.ref = f"A1:{"D"}{last_row}"

createWorkbook.save(filename="formatted_grades.xlsx")
createWorkbook.close()
